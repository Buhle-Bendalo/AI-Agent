import pandas as pd
from datetime import datetime
import os
import io
import json
import pypdf
import openpyxl

# ==========================================
# CONFIGURATION
# ==========================================
MASTER_TEMPLATE_PATH = os.path.join("template", "Master_template.xlsx")
EXAMPLES_DIR = "Examples"
OUTPUT_DIR = "generated_docs"

if not os.path.exists(OUTPUT_DIR):
    os.makedirs(OUTPUT_DIR)


def read_Master_template():
    """Reads the Master Template and returns its structure as text for context."""
    try:
        if not os.path.exists(MASTER_TEMPLATE_PATH):
            return "Error: Master_template.xlsx not found."
        wb = openpyxl.load_workbook(MASTER_TEMPLATE_PATH)
        context = "--- MASTER TEMPLATE STRUCTURE ---\n"
        context += f"Sheets: {wb.sheetnames}\n\n"

        # Front page - show all non-empty cells with coordinates
        ws = wb['Front page']
        context += "FRONT PAGE (key cells to update):\n"
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and not str(cell.value).startswith('='):
                    context += f"  {cell.coordinate}: {repr(cell.value)}\n"

        # QCP - show first 15 rows structure
        ws_qcp = wb['Quality control plan']
        context += "\nQCP PROCESS STEPS (rows 12+, columns A and C):\n"
        for i, row in enumerate(ws_qcp.iter_rows(min_row=11, max_row=30, values_only=True), 11):
            a_val = row[0]
            c_val = row[2] if len(row) > 2 else None
            if a_val or c_val:
                context += f"  Row {i}: A={repr(str(a_val)[:80])}, C={repr(str(c_val)[:80]) if c_val else 'empty'}\n"
        return context
    except Exception as e:
        return f"Error reading template: {str(e)}"


def read_examples():
    """
    Reads the Examples folder to show the agent how previous
    PDF inputs were mapped to Excel outputs.
    """
    if not os.path.exists(EXAMPLES_DIR):
        return "No Examples folder found."

    example_context = "--- REFERENCE EXAMPLES (HOW TO MAP PDF QUOTATION TO EXCEL) ---\n\n"
    files = os.listdir(EXAMPLES_DIR)

    for i in range(1, 5):
        pdf_name = f"input_{i}.pdf"
        xlsx_name = f"output_{i}.xlsx"

        if pdf_name in files and xlsx_name in files:
            example_context += f"[EXAMPLE {i}]\n"

            # Full PDF text
            with open(os.path.join(EXAMPLES_DIR, pdf_name), "rb") as f:
                reader = pypdf.PdfReader(f)
                pdf_text = "".join([p.extract_text() or "" for p in reader.pages])
            example_context += f"INPUT PDF:\n{pdf_text[:2000]}\n\n"

            # Show exact Front page cell values from the output
            wb = openpyxl.load_workbook(os.path.join(EXAMPLES_DIR, xlsx_name))
            ws = wb['Front page']
            example_context += "OUTPUT - FRONT PAGE KEY CELLS:\n"
            key_cells = ['D3', 'A9', 'A12', 'D14', 'D16', 'D18', 'D20', 'D22', 'D24', 'D26', 'D30', 'D32', 'D34']
            for coord in key_cells:
                val = ws[coord].value
                if val:
                    example_context += f"  {coord}: {repr(val)}\n"

            # Show QCP process steps
            if 'Quality control plan' in wb.sheetnames:
                ws_qcp = wb['Quality control plan']
                example_context += "OUTPUT - QCP PROCESS STEPS (format: step_name|description):\n"
                for row in ws_qcp.iter_rows(min_row=11, max_row=50, values_only=True):
                    a_val = row[0]
                    c_val = row[2] if len(row) > 2 else None
                    if a_val and not str(a_val).startswith('='):
                        desc = str(c_val)[:120] if c_val else ""
                        example_context += f"  {a_val}|{desc}\n"

            example_context += "\n"

    return example_context


def process_file_to_text(file_bytes, file_name):
    try:
        if file_name.lower().endswith('.pdf'):
            reader = pypdf.PdfReader(io.BytesIO(file_bytes))
            return "".join([p.extract_text() or "" for p in reader.pages])
        elif file_name.lower().endswith(('.xlsx', '.xls')):
            df = pd.read_excel(io.BytesIO(file_bytes))
            return df.to_string()
        return "Unsupported format."
    except Exception as e:
        return f"Error: {str(e)}"


def _safe_set(ws, row, col, value):
    """Set cell value only if it's not a MergedCell (non-top-left of a merge)."""
    from openpyxl.cell.cell import MergedCell
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        cell.value = value


def generate_excel_workbook(
    client_name: str,
    document_title: str,
    job_number: str,
    part_description: str,
    quantity: str,
    responsible_person: str,
    customer: str,
    quote_number: str,
    surcotec_ref_number: str,
    date_created: str,
    due_date: str,
    qcp_steps: str,
    spray_materials: str = "",
    customer_ref_number: str = "N/A",
    customer_job_number: str = "N/A",
    customer_drawing_number: str = "N/A",
) -> str:
    """
    Generates an Excel workbook by filling the Master Template with the provided
    client job data.

    Parameters:
    - client_name: Company name (used in filename)
    - document_title: Full document title, e.g. 'ST-F-09-01-8111 - Rev1 - Atlantis Foundries - 160O Linear Shafts for HVOF Coating'
    - job_number: Internal job number, e.g. 'IOB 22941'
    - part_description: Description of part, e.g. '160 O Linear Shafts for HVOF Coating'
    - quantity: Number of parts, e.g. '2-off'
    - responsible_person: Surcotec engineer name, e.g. 'Sheldon Deysel'
    - customer: Customer company name, e.g. 'Atlantis Foundries'
    - quote_number: Quote number, e.g. 'SCT - 7725'
    - surcotec_ref_number: Internal ref number, e.g. '26-01-027'
    - date_created: Date in DD.MM.YYYY format, e.g. '20.01.2026'
    - due_date: Due date or working days, e.g. '10-15 Working Days'
    - qcp_steps: Pipe-and-newline separated process steps for QCP.
                 Each line: 'Step Name|Step description text'
                 Example:
                   Incoming inspection|Visually inspect and record dimensions
                   Pre-machining|Set up shaft in lathe, pre-machine to remove damage
                   Grit blasting|Grit blast to Sa2.5 standard
                   Thermal spray|Spray with 1302 material using ARC system
                   Final machining|Machine to final dimensions
                   Final inspection|Dimensional check and crack test
                   Sealing|Seal with anaerobic sealer
                   Delivery|Deliver to customer works
    - spray_materials: Pipe-and-newline separated spray materials.
                 Each line: 'System|Material name|Code'
                 Example:
                   Arc System|ARC Top Coat|1302 (10T)
                   HVOF System|HVOF Powder|3604 (1275 HGB)
    - customer_ref_number: Customer reference number (default N/A)
    - customer_job_number: Customer job number (default N/A)
    - customer_drawing_number: Customer drawing number (default N/A)
    """
    try:
        if not os.path.exists(MASTER_TEMPLATE_PATH):
            return "Error: Master_template.xlsx not found."

        wb = openpyxl.load_workbook(MASTER_TEMPLATE_PATH)

        # ==========================================
        # UPDATE FRONT PAGE
        # ==========================================
        fp = wb['Front page']
        fp['D3'] = document_title
        fp['A9'] = job_number
        fp['A12'] = part_description
        fp['D14'] = quantity
        fp['D16'] = responsible_person
        fp['D18'] = customer
        fp['D20'] = customer_job_number
        fp['D22'] = customer_ref_number
        fp['D24'] = quote_number
        fp['D26'] = surcotec_ref_number
        fp['D30'] = customer_drawing_number
        fp['D32'] = date_created
        fp['D34'] = due_date

        # ==========================================
        # UPDATE QCP PROCESS STEPS
        # ==========================================
        if 'Quality control plan' in wb.sheetnames and qcp_steps.strip():
            qcp = wb['Quality control plan']
            # Clear existing process step rows (12 to 70)
            for row_num in range(12, 71):
                _safe_set(qcp, row_num, 1, None)
                _safe_set(qcp, row_num, 3, None)
                _safe_set(qcp, row_num, 15, None)  # HOLDING POINT col

            steps = [line.strip() for line in qcp_steps.strip().splitlines() if line.strip()]
            current_row = 12
            for step_line in steps:
                parts = step_line.split('|', 1)
                step_name = parts[0].strip()
                step_desc = parts[1].strip() if len(parts) > 1 else ""
                _safe_set(qcp, current_row, 1, step_name)
                _safe_set(qcp, current_row, 3, step_desc)
                current_row += 1

        # ==========================================
        # UPDATE SPRAY AND MACHINE MATERIALS
        # ==========================================
        spray_sheet_name = None
        for name in wb.sheetnames:
            if 'spray' in name.lower() and 'machine' in name.lower():
                spray_sheet_name = name
                break

        if spray_sheet_name and spray_materials.strip():
            sm = wb[spray_sheet_name]
            material_lines = [l.strip() for l in spray_materials.strip().splitlines() if l.strip()]
            # Update material rows starting at row 27
            for idx, mat_line in enumerate(material_lines):
                parts = mat_line.split('|')
                row_num = 27 + idx
                if len(parts) >= 1:
                    _safe_set(sm, row_num, 1, parts[0].strip())
                if len(parts) >= 2:
                    _safe_set(sm, row_num, 4, parts[1].strip())
                if len(parts) >= 3:
                    _safe_set(sm, row_num, 5, parts[2].strip())

        # ==========================================
        # UPDATE LOG SHEET
        # ==========================================
        if 'Log sheet' in wb.sheetnames:
            log = wb['Log sheet']
            log['C2'] = part_description
            log['H2'] = date_created
            log['C3'] = job_number
            log['H3'] = due_date
            log['C5'] = quantity

        # ==========================================
        # UPDATE COVER PAGE
        # ==========================================
        if 'Cover Page' in wb.sheetnames:
            cp = wb['Cover Page']
            cp['C1'] = customer
            # Extract short doc ref (e.g. "ST-F-09-01-8111") from document_title
            short_ref = document_title.split(' ')[0] if document_title else ""
            cp['A3'] = short_ref
            cp['F3'] = f"Quote Ref : {quote_number}"
            cp['A5'] = f"Ref # {surcotec_ref_number}"
            cp['C6'] = f"{part_description} ({quantity})"
            cp['C8'] = responsible_person

        # ==========================================
        # SAVE OUTPUT
        # ==========================================
        safe_name = "".join(c for c in client_name if c.isalnum() or c in (' ', '_')).replace(' ', '_').strip()
        filename = f"Surcotec_{safe_name}_{datetime.now().strftime('%H%M%S')}.xlsx"
        filepath = os.path.join(OUTPUT_DIR, filename)
        wb.save(filepath)
        return f"SUCCESS: Excel workbook '{filename}' has been generated and is ready to download."
    except Exception as e:
        return f"Error generating workbook: {str(e)}"